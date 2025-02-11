import requests
import schedule
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

class CryptoAnalyzer:
    def __init__(self):
        self.base_url = "https://api.coingecko.com/api/v3"
        self.excel_file = "crypto_data_live.xlsx"
        
    def fetch_top_50_crypto(self):
        """Fetch top 50 cryptocurrencies data from CoinGecko API"""
        try:
            endpoint = f"{self.base_url}/coins/markets"
            params = {
                'vs_currency': 'usd',
                'order': 'market_cap_desc',
                'per_page': 50,
                'page': 1,
                'sparkline': False
            }
            response = requests.get(endpoint, params=params)
            response.raise_for_status()
            return response.json()
        except requests.RequestException as e:
            print(f"Error fetching data: {e}")
            return None

    def process_data(self, data):
        """Process the raw cryptocurrency data"""
        if not data:
            return None
        
        processed_data = []
        for coin in data:
            processed_data.append({
                'Name': coin['name'],
                'Symbol': coin['symbol'].upper(),
                'Price (USD)': coin['current_price'],
                'Market Cap': coin['market_cap'],
                '24h Volume': coin['total_volume'],
                '24h Change %': coin.get('price_change_percentage_24h', 0)
            })
        return processed_data

    def analyze_data(self, data):
        """Perform analysis on the cryptocurrency data"""
        if not data:
            return None

        # Sort by market cap (already sorted from API)
        top_5 = data[:5]
        
        # Calculate average price
        total_price = sum(coin['Price (USD)'] for coin in data)
        avg_price = total_price / len(data)
        
        # Find highest and lowest 24h changes
        sorted_by_change = sorted(data, key=lambda x: x['24h Change %'])
        lowest_change = sorted_by_change[0]
        highest_change = sorted_by_change[-1]
        
        return {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'top_5_by_market_cap': top_5,
            'average_price': avg_price,
            'highest_24h_change': highest_change,
            'lowest_24h_change': lowest_change
        }

    def update_excel(self, data, analysis):
        """Update Excel file with latest data and analysis"""
        if not data or not analysis:
            return

        wb = Workbook()
        
        # Live Data Sheet
        ws_data = wb.active
        ws_data.title = "Live Data"
        
        # Headers
        headers = ['Name', 'Symbol', 'Price (USD)', 'Market Cap', '24h Volume', '24h Change %']
        for col, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, coin in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                value = coin[header]
                if isinstance(value, (int, float)):
                    if header == 'Price (USD)':
                        value = f"${value:,.2f}"
                    elif header in ['Market Cap', '24h Volume']:
                        value = f"${value:,.0f}"
                    elif header == '24h Change %':
                        value = f"{value:.2f}%"
                ws_data.cell(row=row, column=col, value=str(value))
        
        # Analysis Sheet
        ws_analysis = wb.create_sheet("Analysis")
        ws_analysis['A1'] = "Analysis Timestamp"
        ws_analysis['B1'] = analysis['timestamp']
        ws_analysis['A2'] = "Average Price (USD)"
        ws_analysis['B2'] = f"${analysis['average_price']:,.2f}"
        ws_analysis['A3'] = "Highest 24h Change"
        ws_analysis['B3'] = f"{analysis['highest_24h_change']['Name']}: {analysis['highest_24h_change']['24h Change %']:.2f}%"
        ws_analysis['A4'] = "Lowest 24h Change"
        ws_analysis['B4'] = f"{analysis['lowest_24h_change']['Name']}: {analysis['lowest_24h_change']['24h Change %']:.2f}%"
        
        # Top 5 Sheet
        ws_top5 = wb.create_sheet("Top 5 by Market Cap")
        for col, header in enumerate(headers, 1):
            ws_top5.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        for row, coin in enumerate(analysis['top_5_by_market_cap'], 2):
            for col, header in enumerate(headers, 1):
                value = coin[header]
                if isinstance(value, (int, float)):
                    if header == 'Price (USD)':
                        value = f"${value:,.2f}"
                    elif header in ['Market Cap', '24h Volume']:
                        value = f"${value:,.0f}"
                    elif header == '24h Change %':
                        value = f"{value:.2f}%"
                ws_top5.cell(row=row, column=col, value=str(value))
        
        try:
            wb.save(self.excel_file)
        except PermissionError:
            print("Error: Please close the Excel file before updating")

    def update_data(self):
        """Main function to update all data"""
        print(f"\nUpdating data at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        raw_data = self.fetch_top_50_crypto()
        if raw_data:
            processed_data = self.process_data(raw_data)
            analysis = self.analyze_data(processed_data)
            self.update_excel(processed_data, analysis)
            print("Data updated successfully!")
        else:
            print("Failed to update data")

def main():
    analyzer = CryptoAnalyzer()
    
    # Initial update
    analyzer.update_data()
    
    # Schedule updates every 5 minutes
    schedule.every(5).minutes.do(analyzer.update_data)
    
    print("\nCrypto Analyzer is running. Press Ctrl+C to stop.")
    print("Excel file will be updated every 5 minutes.")
    
    try:
        while True:
            schedule.run_pending()
            time.sleep(1)
    except KeyboardInterrupt:
        print("\nStopping Crypto Analyzer...")

if __name__ == "__main__":
    main()
